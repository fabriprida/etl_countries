import pandas as pd
from sqlalchemy import text
from sqlalchemy.orm import sessionmaker
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image
import os

def remove_duplicate_languages(languages):
    unique_languages = list(set(languages.split(', ')))
    return ', '.join(unique_languages)

def generate_excel(engine):
    Session = sessionmaker(bind=engine)
    session = Session()

    query1 = """
        SELECT countries.id, countries.name, countries.capital, countries.population, 
               countries.flag, languages.name as language, continents.name as continent, 
               currencies.name as currency
        FROM countries
        LEFT JOIN country_languages ON countries.id = country_languages.country_id
        LEFT JOIN languages ON country_languages.language_id = languages.id
        LEFT JOIN continents ON countries.continent_id = continents.id
        LEFT JOIN country_currencies ON countries.id = country_currencies.country_id
        LEFT JOIN currencies ON country_currencies.currency_id = currencies.id
    """

    df_countries = pd.read_sql_query(query1, engine)

    df_countries['language'] = df_countries['language'].fillna('')

    df_countries['continent'] = df_countries['continent'].fillna('')
    df_countries['currency'] = df_countries['currency'].fillna('')

    df_countries = df_countries.groupby(['id', 'name', 'capital', 'population', 'flag', 'continent']).agg({
        'language': ', '.join,
        'currency': ', '.join
    }).reset_index()
    df_countries['language'] = df_countries['language'].apply(remove_duplicate_languages)

    df_countries = df_countries[['id', 'name', 'capital', 'currency', 'continent', 'language', 'population', 'flag']]
    excel_writer = pd.ExcelWriter('paises_con_metricas.xlsx')
    df_countries.to_excel(excel_writer, sheet_name='Paises', index=False)
    excel_writer._save()
    
    plot_kpis(engine)

def plot_kpis(engine):
    Session = sessionmaker(bind=engine)
    session = Session()

    query = text("""
        SELECT continents.name, COUNT(countries.id) as country_count
        FROM continents
        LEFT JOIN countries ON continents.id = countries.continent_id
        GROUP BY continents.name
    """)

    result = session.execute(query)

    continents = []
    counts = []
    total_countries = 0

    for row in result:
        continent_name, country_count = row
        continents.append(continent_name)
        counts.append(country_count)
        total_countries += country_count

    percentages = [(count / total_countries) * 100 for count in counts]

    plt.figure(figsize=(10, 6))
    plt.pie(percentages, labels=continents, autopct='%1.1f%%', startangle=140)
    plt.title('Percentage of Countries by Continent')
    plt.axis('equal')  
    chart_image_path = 'pie_chart.png'
    plt.savefig(chart_image_path)
    plt.close()

    workbook = load_workbook('paises_con_metricas.xlsx')

    chart_sheet = workbook.create_sheet(title='Metricas')

    chart_img = Image(chart_image_path)
    chart_img.width = 500  
    chart_img.height = 400  
    chart_sheet.add_image(chart_img, 'D2')  

    
    language_query = text("""
        SELECT languages.name, COUNT(countries.id) as country_count
        FROM languages
        LEFT JOIN country_languages ON languages.id = country_languages.language_id
        LEFT JOIN countries ON country_languages.country_id = countries.id
        GROUP BY languages.name
    """)

    language_df = pd.read_sql_query(language_query, engine)
    language_df = language_df.rename(columns={'name': 'language'})

    for r_idx, row in enumerate(dataframe_to_rows(language_df, index=False, header=True), 2):
        for c_idx, value in enumerate(row, 1):
            chart_sheet.cell(row=r_idx, column=c_idx, value=value)

    workbook.save('paises_con_metricas.xlsx')

    os.remove(chart_image_path)
